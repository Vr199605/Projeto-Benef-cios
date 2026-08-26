# -*- coding: utf-8 -*-
"""
etl_reimplantacao.py
=====================
Motor de transferência de dados (PC ELEA / PC PIEMONTE / bases parecidas)
-> Reimplantação ELEA.

Este motor é "schema-adaptive": em vez de depender de nomes fixos de aba e
de coluna na planilha de origem, ele:
  1. varre TODAS as abas do arquivo de origem;
  2. localiza automaticamente a linha de cabeçalho de cada aba (procurando
     pela coluna de certificado nas primeiras linhas);
  3. reconhece cada campo por comparação tolerante a acentos, maiúsculas/
     minúsculas, abreviações e erros de encoding (ex.: "N�MERO" ~ "NUMERO"),
     usando os apelidos configurados em mapping_config.py;
  4. identifica automaticamente qual aba é a de TITULAR (a que tem mais
     linhas com nome do segurado) e qual é a de DEPENDENTES (a que tem
     código de dependente);
  5. junta os dados de todas as abas relacionadas por certificado (titular)
     ou por certificado+código do dependente (dependente);
  6. preenche o template de destino célula a célula com openpyxl,
     preservando 100% da formatação original.

Se uma base de origem trouxer uma coluna que os apelidos configurados não
reconhecem, o campo correspondente fica em branco no destino e um aviso é
registrado no log — o processo NUNCA é interrompido por isso.

Uso via linha de comando:
    python etl_reimplantacao.py --origem PC_PIEMONTE_062026_1.xlsx \
                                 --template Reimplantação_ELEA_EM_BRANCO.xlsx \
                                 --saida Reimplantacao_ELEA_PREENCHIDA.xlsx

Uso programático (também usado pelo app Streamlit):
    from etl_reimplantacao import run_etl
    resumo = run_etl("PC_PIEMONTE_062026_1.xlsx",
                      "Reimplantação_ELEA_EM_BRANCO.xlsx",
                      "saida.xlsx")
"""

from __future__ import annotations

import argparse
import copy
import difflib
import logging
import re
import unicodedata
from datetime import datetime
from typing import Any, Dict, List, Optional

import openpyxl
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

import mapping_config as cfg

# -----------------------------------------------------------------------
# Logging
# -----------------------------------------------------------------------
logger = logging.getLogger("etl_reimplantacao")
if not logger.handlers:
    _h = logging.StreamHandler()
    _h.setFormatter(logging.Formatter("%(asctime)s | %(levelname)-8s | %(message)s"))
    logger.addHandler(_h)
logger.setLevel(logging.INFO)

FUZZY_MATCH_THRESHOLD = 0.82
HEADER_SCAN_ROWS = 8  # quantas linhas iniciais varrer procurando o cabeçalho


# =========================================================================
# 1. NORMALIZAÇÃO DE TEXTO / RECONHECIMENTO TOLERANTE DE CABEÇALHOS
# =========================================================================

def normalize_header(value: Any) -> str:
    """
    Normaliza um texto de cabeçalho para comparação tolerante:
    - remove acentos
    - remove o caractere de substituição Unicode '\\ufffd' (comum em
      arquivos com problema de encoding na origem)
    - maiúsculas, sem pontuação, espaços únicos
    """
    if value is None:
        return ""
    text = str(value)
    text = text.replace("\ufffd", "")
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.upper()
    text = re.sub(r"[^A-Z0-9 ]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _best_alias_match(header_norm: str, aliases_norm: List[str]) -> float:
    """Maior score de similaridade entre um cabeçalho e uma lista de apelidos."""
    if not header_norm:
        return 0.0
    best = 0.0
    for alias in aliases_norm:
        # match exato ou "contém" pontuam alto automaticamente
        if header_norm == alias:
            return 1.0
        if alias in header_norm or header_norm in alias:
            best = max(best, 0.95)
            continue
        ratio = difflib.SequenceMatcher(None, header_norm, alias).ratio()
        best = max(best, ratio)
    return best


def match_field_to_columns(
    field_aliases: Dict[str, List[str]],
    normalized_columns: List[str],
) -> Dict[str, Optional[str]]:
    """
    Para cada campo canônico, encontra a coluna (já normalizada) da aba que
    melhor corresponde a algum de seus apelidos, respeitando a ordem de
    prioridade dos apelidos (uma versão "(Y2K)", por exemplo, vence uma
    versão comum caso ambas estejam presentes).
    Retorna {campo_canonico: nome_coluna_normalizado ou None}.
    """
    result: Dict[str, Optional[str]] = {}
    for field, aliases in field_aliases.items():
        aliases_norm = [normalize_header(a) for a in aliases]
        best_col, best_score, best_alias_rank = None, 0.0, None
        for col in normalized_columns:
            for rank, alias in enumerate(aliases_norm):
                score = _best_alias_match(col, [alias])
                if score >= FUZZY_MATCH_THRESHOLD:
                    # prioridade: primeiro por rank do apelido (menor = melhor),
                    # depois por score de similaridade
                    if (
                        best_alias_rank is None
                        or rank < best_alias_rank
                        or (rank == best_alias_rank and score > best_score)
                    ):
                        best_col, best_score, best_alias_rank = col, score, rank
        result[field] = best_col
    return result


# =========================================================================
# 2. FUNÇÕES DE TRATAMENTO / CONVERSÃO DE DADOS
# =========================================================================

def _only_digits(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    return re.sub(r"\D", "", str(value))


def clean_text(value: Any) -> Optional[str]:
    """strip() em textos; NaN/None/"" viram None (célula em branco)."""
    if value is None:
        return None
    if isinstance(value, float) and pd.isna(value):
        return None
    text = str(value).strip()
    return text if text else None


NULL_DATE_PLACEHOLDERS = {"00/00/0000", "01/01/0001", "00/00/00"}


def data_ddmmyyyy(value: Any) -> Optional[int]:
    """
    Converte uma data textual para inteiro DDMMYYYY, compatível com o
    number_format '00000000' do template (ex.: 02/09/1957 -> 2091957).

    Suporta ano com 4 e com 2 dígitos ('14/09/90'); para ano de 2 dígitos
    usa a regra: 00-49 -> 20XX, 50-99 -> 19XX (comum em arquivos legados).
    Placeholders de data nula (ex.: 00/00/0000) retornam None.
    """
    text = clean_text(value)
    if text is None or text in NULL_DATE_PLACEHOLDERS:
        return None

    dt = None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            dt = datetime.strptime(text, fmt)
            break
        except ValueError:
            continue

    if dt is None:
        # tenta formato com ano de 2 dígitos: dd/mm/yy
        m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{2})$", text)
        if m:
            day, month, yy = int(m.group(1)), int(m.group(2)), int(m.group(3))
            year = 2000 + yy if yy < 50 else 1900 + yy
            try:
                dt = datetime(year, month, day)
            except ValueError:
                dt = None

    if dt is None and isinstance(value, datetime):
        dt = value

    if dt is None or dt.year < 1900:
        return None
    return int(dt.strftime("%d%m%Y"))


def cpf_base(value: Any) -> Optional[str]:
    digits = _only_digits(value)
    if not digits:
        return None
    digits = digits.zfill(11)[-11:]
    return digits[:9]


def cpf_control(value: Any) -> Optional[str]:
    digits = _only_digits(value)
    if not digits:
        return None
    digits = digits.zfill(11)[-11:]
    return digits[9:11]


def cep_base(value: Any) -> Optional[str]:
    digits = _only_digits(value)
    if not digits:
        return None
    digits = digits.zfill(8)[-8:]
    return digits[:5]


def cep_comp(value: Any) -> Optional[str]:
    digits = _only_digits(value)
    if not digits:
        return None
    digits = digits.zfill(8)[-8:]
    return digits[5:8]


def sexo_code(value: Any) -> Optional[int]:
    text = clean_text(value)
    if text is None:
        return None
    return cfg.SEXO_MAP.get(text.upper())


def estado_civil_code(value: Any) -> Optional[int]:
    text = clean_text(value)
    if text is None:
        return None
    return cfg.ESTADO_CIVIL_MAP.get(text.upper())


def sim_nao(value: Any) -> Optional[str]:
    text = clean_text(value)
    if text is None:
        return None
    text = text.upper()
    if text in ("S", "SIM", "Y", "YES", "1", "TRUE"):
        return "S"
    if text in ("N", "NAO", "NÃO", "0", "FALSE"):
        return "N"
    return text[:1]


def _telefone_com_ddd(record: Dict[str, Any]) -> Optional[str]:
    ddd = _only_digits(record.get("DDD1"))
    fone = _only_digits(record.get("TELEFONE1"))
    if not fone:
        return None
    return f"{ddd}{fone}" if ddd else fone


SPECIAL_HANDLERS = {
    "cpf_base": lambda record, raw: cpf_base(raw),
    "cpf_control": lambda record, raw: cpf_control(raw),
    "cep_base": lambda record, raw: cep_base(raw),
    "cep_comp": lambda record, raw: cep_comp(raw),
    "data_ddmmyyyy": lambda record, raw: data_ddmmyyyy(raw),
    "sexo_code": lambda record, raw: sexo_code(raw),
    "estado_civil_code": lambda record, raw: estado_civil_code(raw),
    "sim_nao": lambda record, raw: sim_nao(raw),
    "telefone_com_ddd": lambda record, raw: _telefone_com_ddd(record),
}


# =========================================================================
# 3. LEITURA ADAPTATIVA DA BASE DE ORIGEM
# =========================================================================

class SourceSheet:
    """Representa uma aba de origem já lida, com cabeçalho normalizado."""

    def __init__(self, name: str, header_row: int, df: pd.DataFrame, columns_norm: List[str]):
        self.name = name
        self.header_row = header_row
        self.df = df  # colunas = nomes ORIGINAIS (não normalizados)
        self.columns_norm = columns_norm  # paralelo às colunas originais


def _find_header_row(ws: Worksheet, key_aliases_norm: List[str]) -> int:
    """
    Procura, nas primeiras HEADER_SCAN_ROWS linhas, aquela cujo conjunto de
    valores contém uma célula compatível com algum apelido de "certificado"
    (ou outro campo-chave). Se não encontrar, usa a linha 1 como fallback.
    """
    max_col = min(ws.max_column, 40)
    for r in range(1, HEADER_SCAN_ROWS + 1):
        row_values = [normalize_header(ws.cell(row=r, column=c).value) for c in range(1, max_col + 1)]
        for val in row_values:
            if not val:
                continue
            for alias in key_aliases_norm:
                if _best_alias_match(val, [alias]) >= FUZZY_MATCH_THRESHOLD:
                    return r
    return 1


def load_source_workbook(path: str) -> List[SourceSheet]:
    """Lê todas as abas do arquivo de origem, detectando o cabeçalho de cada uma."""
    wb = openpyxl.load_workbook(path, data_only=True)
    key_aliases_norm = [normalize_header(a) for a in cfg.TITULAR_FIELD_ALIASES["CERTIFICADO"]]

    sheets: List[SourceSheet] = []
    for ws in wb.worksheets:
        header_row = _find_header_row(ws, key_aliases_norm)
        try:
            df = pd.read_excel(path, sheet_name=ws.title, header=header_row - 1, dtype=str)
        except Exception:
            logger.exception("Falha ao ler a aba '%s' como tabela — pulando.", ws.title)
            continue
        df.columns = [str(c).strip() for c in df.columns]
        df = df.dropna(how="all")
        columns_norm = [normalize_header(c) for c in df.columns]
        sheets.append(SourceSheet(ws.title, header_row, df, columns_norm))
        logger.info(
            "Origem: aba '%s' lida (cabeçalho na linha %d) com %d linha(s) de dado.",
            ws.title, header_row, len(df),
        )
    return sheets


def _sheet_has_field(sheet: SourceSheet, field_aliases: List[str]) -> bool:
    aliases_norm = [normalize_header(a) for a in field_aliases]
    for col in sheet.columns_norm:
        if _best_alias_match(col, aliases_norm) >= FUZZY_MATCH_THRESHOLD:
            return True
    return False


def _original_column_for(sheet: SourceSheet, normalized_col_name: str) -> Optional[str]:
    for orig_col, norm_col in zip(sheet.df.columns, sheet.columns_norm):
        if norm_col == normalized_col_name:
            return orig_col
    return None


def build_records(
    sheets: List[SourceSheet],
    field_aliases: Dict[str, List[str]],
    key_field: str,
    role_hint_field: str,
    exclude_sheets: Optional[List[str]] = None,
) -> List[Dict[str, Any]]:
    """
    Identifica quais abas participam desta "família" de registros (contêm a
    coluna-chave) e monta uma lista de registros unificados, na ordem da
    aba "principal" (a que tem mais linhas entre as candidatas — ou a que
    contém o campo-dica role_hint_field, em caso de empate).

    Cada registro é um dicionário {campo_canonico: valor}, resolvido pela
    primeira aba (na ordem de prioridade) que tiver aquele campo preenchido.
    """
    exclude_sheets = exclude_sheets or []
    key_aliases = field_aliases[key_field]

    candidate_sheets = [
        s for s in sheets
        if s.name not in exclude_sheets and _sheet_has_field(s, key_aliases) and len(s.df) > 0
    ]

    if not candidate_sheets:
        logger.warning(
            "Nenhuma aba de origem contém um campo compatível com '%s' — "
            "esta família de registros ficará vazia.", key_field,
        )
        return []

    # aba principal = a que tem mais linhas (empiricamente, a tabela "mestre")
    primary_sheet = max(candidate_sheets, key=lambda s: len(s.df))
    other_sheets = [s for s in candidate_sheets if s is not primary_sheet]

    # Para cada aba candidata, resolve o mapeamento campo -> coluna original,
    # e monta um índice por chave para lookup rápido.
    sheet_field_map: Dict[str, Dict[str, Optional[str]]] = {}
    sheet_index: Dict[str, pd.DataFrame] = {}
    for s in candidate_sheets:
        resolved = match_field_to_columns(field_aliases, s.columns_norm)
        sheet_field_map[s.name] = resolved
        key_col = resolved.get(key_field)
        key_orig = _original_column_for(s, key_col) if key_col else None
        if key_orig is None:
            continue
        df = s.df.copy()
        df["_key_norm"] = df[key_orig].astype(str).str.strip().str.lstrip("0")
        sheet_index[s.name] = df

    unmatched_fields = [f for f, col in sheet_field_map[primary_sheet.name].items() if col is None]
    if unmatched_fields:
        logger.warning(
            "Aba '%s': os seguintes campos não foram localizados nos "
            "cabeçalhos e ficarão em branco para os registros desta aba: %s",
            primary_sheet.name, ", ".join(unmatched_fields),
        )

    records: List[Dict[str, Any]] = []
    primary_key_col = sheet_field_map[primary_sheet.name].get(key_field)
    primary_key_orig = _original_column_for(primary_sheet, primary_key_col) if primary_key_col else None

    if primary_key_orig is None:
        logger.error(
            "Aba principal '%s' não possui coluna-chave reconhecida — "
            "esta família de registros ficará vazia.", primary_sheet.name,
        )
        return []

    for _, row in primary_sheet.df.iterrows():
        key_val = str(row[primary_key_orig]).strip()
        if not key_val or key_val.lower() == "nan":
            continue
        key_norm = key_val.lstrip("0")

        record: Dict[str, Any] = {}
        # 1) preenche a partir da aba principal
        for field, norm_col in sheet_field_map[primary_sheet.name].items():
            orig_col = _original_column_for(primary_sheet, norm_col) if norm_col else None
            record[field] = row[orig_col] if orig_col else None

        # 2) completa (só onde ainda está vazio) a partir das demais abas
        for s in other_sheets:
            if s.name not in sheet_index:
                continue
            match = sheet_index[s.name][sheet_index[s.name]["_key_norm"] == key_norm]
            if match.empty:
                continue
            match_row = match.iloc[0]
            for field, norm_col in sheet_field_map[s.name].items():
                if record.get(field) not in (None, ""):
                    continue
                orig_col = _original_column_for(s, norm_col) if norm_col else None
                if orig_col:
                    record[field] = match_row[orig_col]

        records.append(record)

    logger.info(
        "Família '%s': %d registro(s) unificado(s) a partir da aba principal '%s' "
        "(+ %d aba(s) complementares: %s).",
        key_field, len(records), primary_sheet.name, len(other_sheets),
        ", ".join(s.name for s in other_sheets) or "nenhuma",
    )
    return records


def build_titular_records(sheets: List[SourceSheet]) -> List[Dict[str, Any]]:
    # Exclui abas que sejam claramente de dependente, para não misturar.
    dep_alias = cfg.DEPENDENTE_FIELD_ALIASES[cfg.DEPENDENTE_ROLE_HINT_FIELD]
    dep_sheet_names = [s.name for s in sheets if _sheet_has_field(s, dep_alias)]
    return build_records(
        sheets,
        cfg.TITULAR_FIELD_ALIASES,
        cfg.TITULAR_KEY_FIELD,
        cfg.TITULAR_PRIMARY_HINT_FIELD,
        exclude_sheets=dep_sheet_names,
    )


def build_dependente_records(sheets: List[SourceSheet]) -> List[Dict[str, Any]]:
    dep_alias = cfg.DEPENDENTE_FIELD_ALIASES[cfg.DEPENDENTE_ROLE_HINT_FIELD]
    dep_sheet_names = [s.name for s in sheets if _sheet_has_field(s, dep_alias)]
    other_sheets = [s.name for s in sheets if s.name not in dep_sheet_names]
    return build_records(
        sheets,
        cfg.DEPENDENTE_FIELD_ALIASES,
        "CERTIFICADO",
        cfg.DEPENDENTE_ROLE_HINT_FIELD,
        exclude_sheets=other_sheets,
    )


# =========================================================================
# 4. ESCRITA NO TEMPLATE DE DESTINO (openpyxl, célula a célula)
# =========================================================================

def _header_map(ws: Worksheet, header_row: int) -> Dict[str, int]:
    mapping = {}
    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(row=header_row, column=col_idx).value
        if header is not None:
            mapping[str(header).strip()] = col_idx
    return mapping


def _copy_style(src_cell, dst_cell) -> None:
    dst_cell.font = copy.copy(src_cell.font)
    dst_cell.fill = copy.copy(src_cell.fill)
    dst_cell.border = copy.copy(src_cell.border)
    dst_cell.alignment = copy.copy(src_cell.alignment)
    dst_cell.number_format = src_cell.number_format
    dst_cell.protection = copy.copy(src_cell.protection)


def _resolve_value(record: Dict[str, Any], sheet_name: str, dst_col: str, canonical_field: str) -> Any:
    handler_name = cfg.SPECIAL_COLUMNS.get(sheet_name, {}).get(dst_col)
    raw = record.get(canonical_field)
    if handler_name:
        handler = SPECIAL_HANDLERS[handler_name]
        try:
            return handler(record, raw)
        except Exception:
            logger.warning(
                "Falha ao converter valor '%r' (campo '%s', aba '%s', handler '%s') "
                "— célula ficará em branco.", raw, canonical_field, sheet_name, handler_name,
            )
            return None
    return clean_text(raw)


def fill_sheet(ws: Worksheet, sheet_name: str, records: List[Dict[str, Any]]) -> int:
    """Preenche uma aba do template com os registros, célula a célula."""
    columns_map = cfg.SHEET_CONFIG[sheet_name]["columns"]
    header_idx = _header_map(ws, cfg.DST_HEADER_ROW)
    data_start = cfg.DST_DATA_START_ROW
    template_row = cfg.DST_TEMPLATE_ROW
    max_col = ws.max_column

    template_snapshot = []
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=template_row, column=col_idx)
        template_snapshot.append((cell.value, cell))

    original_last_row = ws.max_row
    rows_with_error = 0

    for i, record in enumerate(records):
        dest_row = data_start + i
        try:
            if dest_row != template_row:
                for col_idx in range(1, max_col + 1):
                    tmpl_value, tmpl_cell = template_snapshot[col_idx - 1]
                    dst_cell = ws.cell(row=dest_row, column=col_idx)
                    _copy_style(tmpl_cell, dst_cell)
                    dst_cell.value = tmpl_value

            for dst_col, canonical_field in columns_map.items():
                if dst_col not in header_idx:
                    logger.warning(
                        "Aba '%s': coluna de destino '%s' não encontrada no "
                        "cabeçalho do template — verifique mapping_config.py.",
                        sheet_name, dst_col,
                    )
                    continue
                value = _resolve_value(record, sheet_name, dst_col, canonical_field)
                ws.cell(row=dest_row, column=header_idx[dst_col]).value = value
        except Exception:
            rows_with_error += 1
            logger.exception(
                "Aba '%s': erro ao gravar o registro #%d (linha %d) — linha "
                "pulada, processamento continua.", sheet_name, i + 1, dest_row,
            )

    last_needed_row = data_start + len(records) - 1
    for row_idx in range(last_needed_row + 1, original_last_row + 1):
        for col_idx in range(1, max_col + 1):
            ws.cell(row=row_idx, column=col_idx).value = None

    if rows_with_error:
        logger.warning("Aba '%s': %d registro(s) com erro (pulados).", sheet_name, rows_with_error)

    return len(records) - rows_with_error


# =========================================================================
# 5. ORQUESTRAÇÃO
# =========================================================================

def run_etl(source_path: str, template_path: str, output_path: str) -> Dict[str, int]:
    """Executa o processo completo. Retorna {aba: linhas_processadas}."""
    summary: Dict[str, int] = {}

    logger.info("Lendo base de origem: %s", source_path)
    sheets = load_source_workbook(source_path)
    if not sheets:
        raise ValueError("Não foi possível ler nenhuma aba da planilha de origem.")

    logger.info("Identificando e unificando registros de TITULAR...")
    titular_records = build_titular_records(sheets)
    logger.info("Total de titulares unificados: %d", len(titular_records))

    logger.info("Identificando e unificando registros de DEPENDENTE...")
    dependente_records = build_dependente_records(sheets)
    logger.info("Total de dependentes unificados: %d", len(dependente_records))

    families = {"titular": titular_records, "dependente": dependente_records}

    logger.info("Abrindo template de destino: %s", template_path)
    wb = load_workbook(template_path)

    for sheet_name in cfg.DST_SHEETS:
        try:
            if sheet_name not in wb.sheetnames:
                logger.error("Aba '%s' não encontrada no template — pulando.", sheet_name)
                continue
            ws = wb[sheet_name]
            family = cfg.SHEET_CONFIG[sheet_name]["family"]
            records = families[family]
            n = fill_sheet(ws, sheet_name, records)
            summary[sheet_name] = n
            logger.info("Aba '%s': %d linha(s) processada(s) com sucesso.", sheet_name, n)
        except Exception:
            logger.exception("Falha ao processar a aba '%s'.", sheet_name)
            summary[sheet_name] = 0

    logger.info("Salvando arquivo final: %s", output_path)
    wb.save(output_path)
    logger.info("Concluído.")
    return summary


# =========================================================================
# 6. CLI
# =========================================================================

def _parse_args():
    parser = argparse.ArgumentParser(description="ETL bases de posição cadastral -> Reimplantação ELEA")
    parser.add_argument("--origem", required=True, help="Caminho do arquivo de origem (base bruta)")
    parser.add_argument("--template", required=True, help="Caminho do template em branco (destino)")
    parser.add_argument("--saida", required=True, help="Caminho do arquivo .xlsx a ser gerado")
    return parser.parse_args()


if __name__ == "__main__":
    args = _parse_args()
    resumo = run_etl(args.origem, args.template, args.saida)
    print("\nResumo do processamento:")
    for aba, qtd in resumo.items():
        print(f"  - {aba}: {qtd} linha(s)")
